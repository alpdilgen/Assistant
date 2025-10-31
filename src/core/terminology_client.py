"""Integration helpers for delegating terminology extraction to Termextractor."""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import List, Mapping, Optional, Sequence, TYPE_CHECKING
from urllib.parse import urlencode
from xml.etree import ElementTree as ET

import streamlit as st
from openpyxl import load_workbook

if TYPE_CHECKING:  # pragma: no cover - typing only
    from streamlit.runtime.uploaded_file_manager import UploadedFile

DEFAULT_TERMEXTRACTOR_URL = "https://termtool.streamlit.app"


def send_to_external_termextractor(
    files: Sequence[io.BufferedIOBase],
    src_lang: str,
    tgt_lang: str,
    project_name: Optional[str] = None,
    *,
    base_url: Optional[str] = None,
) -> str:
    """Render a link guiding the user to the hosted Termextractor application."""

    params: dict[str, str] = {
        "source_language": src_lang,
        "target_language": tgt_lang,
    }
    if project_name:
        params["project_name"] = project_name

    query = urlencode(params)
    base = base_url or DEFAULT_TERMEXTRACTOR_URL
    url = f"{base}?{query}" if query else base

    st.info(
        "Files will be processed by the external Termextractor. Open the tool in a new tab to upload the same files."
    )

    if hasattr(st, "link_button"):
        st.link_button("Go to Termextractor", url)
    else:  # pragma: no cover - fallback for older Streamlit versions
        st.markdown(f"[Go to Termextractor]({url})")

    if files:
        st.caption("Files prepared for terminology hand-off:")
        for file_obj in files:
            st.write(f"• {getattr(file_obj, 'name', 'uploaded_file')}")

    return url


def _normalise_row(row: Mapping[str, str]) -> dict:
    lower = {key.lower(): value for key, value in row.items()}
    term = lower.get("term") or lower.get("source term") or lower.get("source")
    translation = lower.get("translation") or lower.get("target term") or lower.get("target")
    notes = lower.get("notes") or lower.get("comment") or lower.get("context")
    category = lower.get("category") or lower.get("domain") or ""
    dnt_flag = str(lower.get("dnt") or lower.get("do not translate") or "").lower()
    dnt = dnt_flag in {"yes", "y", "true", "1", "dnt"}
    return {
        "term": term or "",
        "translation": translation or "",
        "notes": notes or "",
        "category": category or "",
        "dnt": dnt,
    }


def _parse_csv(data: bytes) -> List[dict]:
    text = data.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [_normalise_row(row) for row in reader]


def _parse_xlsx(data: bytes) -> List[dict]:
    buffer = io.BytesIO(data)
    workbook = load_workbook(buffer, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    entries: List[dict] = []
    for values in rows[1:]:
        row = {headers[idx].lower(): str(value).strip() if value is not None else "" for idx, value in enumerate(values)}
        entries.append(_normalise_row(row))
    return entries


def _parse_tbx(data: bytes) -> List[dict]:
    entries: List[dict] = []
    root = ET.fromstring(data)
    for term_entry in root.findall(".//termEntry"):
        lang_sets = term_entry.findall("langSet")
        term_text = ""
        translation = ""
        for lang_set in lang_sets:
            lang = lang_set.get("{http://www.w3.org/XML/1998/namespace}lang", "")
            term = lang_set.findtext(".//term") or ""
            if not term_text:
                term_text = term
            else:
                translation = translation or term
        entries.append(
            {
                "term": term_text,
                "translation": translation,
                "notes": term_entry.findtext("descrip") or "",
                "category": term_entry.get("subjectField", ""),
                "dnt": False,
            }
        )
    return entries


def parse_terminology_file(upload: "UploadedFile") -> List[dict]:
    """Parse terminology output files (CSV/XLSX/TBX) into a unified structure."""

    data = upload.getvalue()
    extension = Path(upload.name).suffix.lower()
    if extension == ".csv":
        return _parse_csv(data)
    if extension in {".xlsx", ".xlsm"}:
        return _parse_xlsx(data)
    if extension == ".tbx":
        return _parse_tbx(data)
    raise ValueError(f"Unsupported terminology format: {extension}")


__all__ = ["parse_terminology_file", "send_to_external_termextractor"]
